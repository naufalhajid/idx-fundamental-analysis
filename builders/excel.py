import openpyxl

from builders.analysers import StockPriceAnalyser
from builders.analysers.fundamental_analyser import FundamentalAnalyser
from builders.analysers.key_analysis_analyser import KeyAnalysisAnalyser
from builders.analysers.sentiment_analyser import SentimentAnalyser
from builders.builder_interface import BuilderInterface
from utils.logger_config import logger


class Excel(BuilderInterface):
    def __init__(
        self,
        title: str,
        fundamental_analyser: FundamentalAnalyser,
        sentiment_analyser: SentimentAnalyser,
        key_analysis_analyser: KeyAnalysisAnalyser,
        stock_price_analyser: StockPriceAnalyser,
    ):
        self.filename = f"./output/{title}.xlsx"
        self.fundamental_analyser = fundamental_analyser
        self.sentiment_analyser = sentiment_analyser
        self.key_analysis_analyser = key_analysis_analyser
        self.stock_price_analyser = stock_price_analyser

        try:
            # Try to load an existing workbook
            self.wb = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            # Create a new workbook if it doesn't exist
            self.wb = openpyxl.Workbook()

            # Remove the default sheet created.
            self.wb.remove(self.wb.active)

    def _write_to_sheet(self, sheet_name: str, values: []):
        """
        Write values to an existing or new sheet in the Excel file.

        :param sheet_name: Name of the sheet
        :param values: List of rows (each row is a list of values)
        """
        # Create a new sheet if it doesn't exist
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(title=sheet_name)

        sheet = self.wb[sheet_name]

        for i, row_data in enumerate(values):
            for j, value in enumerate(row_data):
                sheet.cell(row=i + 1, column=j + 1, value=value)

    def save(self):
        """
        Save the workbook to a file.
        """
        self.wb.save(self.filename)
        logger.info(
            f"Excel file saved successfully in the root project (./{self.filename})"
        )

    def insert_stock(self):
        """
        Inserts stock data into the spreadsheet.
        """
        self._write_to_sheet("idx-stocks", self.fundamental_analyser.stocks_sheet())

    def insert_key_statistic(self):
        """
        Inserts keystatistic data into the spreadsheet.
        """
        self._write_to_sheet(
            "key-statistics", self.fundamental_analyser.key_statistics_sheet()
        )

    def insert_key_analysis(self):
        """
        Inserts fundamental analysis data into the spreadsheet.
        """

        self._write_to_sheet("analysis", self.key_analysis_analyser.analysis_sheet())

    def insert_sentiment(self):
        """
        Inserts sentiment analysis data into the spreadsheet.
        """
        self._write_to_sheet("sentiments", self.sentiment_analyser.sentiment_sheet())

    def insert_stock_price(self):
        """
        Inserts stock price data into the spreadsheet.
        """
        self._write_to_sheet(
            "stock-prices", self.stock_price_analyser.stock_price_sheet()
        )
